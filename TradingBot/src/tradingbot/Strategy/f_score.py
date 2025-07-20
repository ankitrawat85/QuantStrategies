import yfinance as yf
import pandas as pd

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
import sys
from collections import defaultdict
from typing import Dict, List

from tradingbot.utils.functions.commonFunctions import trim_dataframe  , STANDARD_NAMES,get_standard_name,standardize_column_names,F_SCORE_REQUIREMENTS

# Define F-Score columns
F_SCORE_COLUMNS = [
    'f_score',
    'profitability_score',
    'leverage_liquidity_score',
    'operating_efficiency_score'
]

def safe_divide(numerator, denominator):
    """Safe division with zero handling"""
    return numerator / denominator if denominator else 0

def validate_data(df: pd.DataFrame) -> Dict[str, List[str]]:
    """
    Validate data for F-Score calculation
    Returns dict with validation issues
    """
    issues = defaultdict(list)
    available_cols = set(df.columns)
    
    # Check required columns
    for metric in F_SCORE_REQUIREMENTS['required']:
        if metric not in available_cols:
            issues['missing_required'].append(metric)
    
    # Check data quality
    for col in df.columns:
        if df[col].isnull().any():
            issues['data_quality'].append(f"{col} has null values")
        if (df[col] == 0).any():
            issues['data_quality'].append(f"{col} has zero values")
    
    return dict(issues)

def read_data_with_custom_index(file_path, sheet_name=None, 
                              datetime_cols=None, index_cols=None, 
                              file_type='excel',composite_key=None):
    """
    Read data file with standardized column names and custom indexing
    """
    # Read file
    if file_type == 'csv':
        df = pd.read_csv(file_path, sheet_name=sheet_name)
    else:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # Standardize column names
    df = standardize_column_names(df)
    
    # Process datetime columns
    if datetime_cols:
        for col, dtype in datetime_cols.items():
            std_col = get_standard_name(col)
            if std_col in df.columns:
                try:
                    if dtype == 'year':
                        df[std_col] = pd.to_datetime(df[std_col], format='%Y').dt.year
                    elif dtype == 'yearquarter':
                        df[std_col] = pd.to_datetime(df[std_col])
                        df['Year'] = df[std_col].dt.year
                        df['Quarter'] = 'Q' + df[std_col].dt.quarter.astype(str)
                    elif dtype == 'date':
                        df[std_col] = pd.to_datetime(df[std_col])
                except Exception as e:
                    print(f"Warning: Could not convert {std_col} to {dtype}: {str(e)}")
    
    # Create MultiIndex if specified
    if index_cols:
        missing_cols = [col for col in index_cols if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Index columns not found: {missing_cols}")
        
        df.index = pd.MultiIndex.from_arrays(
            [df[col] for col in index_cols],
            names=index_cols
        )
        df = df.drop(columns=index_cols)
    
    # Handle composite key by adding it as the first level of index
    if composite_key:
        missing_keys = [key for key in composite_key if key not in df.columns]
        if missing_keys:
            raise ValueError(f"Composite key columns not found: {missing_keys}")
        
        # Add composite key columns to index while preserving existing index
        if isinstance(df.index, pd.MultiIndex):
            df.index = pd.MultiIndex.from_arrays(
                [df[col] for col in composite_key] + [df.index.get_level_values(i) for i in range(len(df.index.levels))],
                names=composite_key + df.index.names
            )
        else:
            df.index = pd.MultiIndex.from_arrays(
                [df[col] for col in composite_key] + [df.index],
                names=composite_key + [df.index.name if df.index.name else 'index']
            )
        df = df.drop(columns=composite_key)
    
    return df

def get_gross_profit(row: pd.Series) -> float:
    """
    Calculate gross profit from row data using available metrics
    Tries multiple calculation methods in order of preference
    """
    # Try direct gross_profit first
    if 'gross_profit' in row.index:
        return row['gross_profit']
    
    # Calculate from revenue and COGS if available
    if all(m in row.index for m in ['revenue', 'cost_of_goods_sold']):
        return row['revenue'] - row['cost_of_goods_sold']
    
    # Calculate from margin percentage if available
    if all(m in row.index for m in ['revenue', 'gross_margin_pct']):
        return row['revenue'] * (row['gross_margin_pct'] / 100)
    
    # If no calculation possible, return NaN which will fail safe_divide
    return np.nan

def safe_divide(numerator: float, denominator: float) -> float:
    """Safe division with zero handling"""
    return numerator / denominator if denominator else 0

def calculate_f_scores(df: pd.DataFrame, composite_key=None) -> pd.DataFrame:
    """
    Calculate Piotroski F-Scores using standardized column names
    Returns original DataFrame with added score columns
    """
    # Validate input data
    for dropCol  in ['index','Unnamed']:
        if dropCol in df.columns:
            df =df.drop(columns=[dropCol])


    validation = validate_data(df)
    if validation.get('missing_required'):
        print(
            f"Missing required columns: {validation['missing_required']}\n"
            f"Available columns: {list(df.columns)}"
        )
    df.reset_index(inplace =True)

    # Create copy for results (preserve original data)
    result_df = df.copy()
    
    # Initialize score columns
    for col in F_SCORE_COLUMNS:
        result_df[col] = np.nan
    
    # Prepare composite key handling
    if composite_key is None:
        composite_key = []
    elif isinstance(composite_key, str):
        composite_key = [composite_key]
    
    # Ensure Year is available
    if 'Year' not in df.columns:
        raise ValueError("'Year' column is required for sorting")
    
    # Create group identifier if composite key exists
    if composite_key:
        # Create temporary group identifier column
        df['_temp_group_'] = df[composite_key].apply(tuple, axis=1)
        groups = df['_temp_group_'].unique()
    else:
        # Treat entire dataframe as one group
        df['_temp_group_'] = None
        groups = [None]
    
    # Process each group separately
    for group in groups:
        # Filter group data
        if composite_key:
            group_mask = df['_temp_group_'] == group
        else:
            group_mask = pd.Series(True, index=df.index)
        
        group_df = df[group_mask].copy()
        
        # Sort by Year within group
        group_df = group_df.sort_values('Year', ascending=True)
        
        # Need at least 2 periods to compare
        if len(group_df) < 2:
            print(f"Skipping group {group} with only {len(group_df)} periods")
            continue
        
        # Calculate scores for each consecutive period
        for i in range(1, len(group_df)):
            current = group_df.iloc[i]
            previous = group_df.iloc[i-1]
            
            try:
                # Calculate scores
                scores = {
                    'f_score': 0,
                    'profitability_score': 0,
                    'leverage_liquidity_score': 0,
                    'operating_efficiency_score': 0
                }
                
                # 1. Profitability Metrics
                roa = safe_divide(current.get('net_income', 0), current.get('total_assets', 1))
                cfo = safe_divide(current.get('total_operating_cash_flow', 0), current.get('total_assets', 1))
                delta_roa = roa - safe_divide(previous.get('net_income', 0), previous.get('total_assets', 1))
                
                scores['profitability_score'] = sum([
                    roa > 0,
                    cfo > 0,
                    delta_roa > 0,
                    cfo > roa
                ])
                
                # 2. Leverage/Liquidity Metrics
                current_ratio = safe_divide(current.get('current_assets', 0), current.get('current_liabilities', 1))
                prev_ratio = safe_divide(previous.get('current_assets', 0), previous.get('current_liabilities', 1))
                
                scores['leverage_liquidity_score'] = sum([
                    current.get('debt_to_equity', 0) < previous.get('debt_to_equity', 0),
                    current_ratio > prev_ratio,
                    current.get('shares_outstanding', 0) <= previous.get('shares_outstanding', 0)
                ])
                
                # 3. Operating Efficiency Metrics
                current_gross = get_gross_profit(current)
                prev_gross = get_gross_profit(previous)
                gross_margin = safe_divide(current_gross, current.get('revenue', 1))
                prev_margin = safe_divide(prev_gross, previous.get('revenue', 1))
                
                scores['operating_efficiency_score'] = sum([
                    gross_margin > prev_margin,
                    safe_divide(current.get('revenue', 0), current.get('total_assets', 1)) > 
                    safe_divide(previous.get('revenue', 0), previous.get('total_assets', 1))
                ])
                
                scores['f_score'] = (
                    scores['profitability_score'] + 
                    scores['leverage_liquidity_score'] + 
                    scores['operating_efficiency_score']
                )
                
                
                
                # Find matching row in result_df to assign scores
                mask = (result_df['Year'] == current['Year'])
                if composite_key:
                    for key_col, key_val in zip(composite_key, group):
                        mask &= (result_df[key_col] == key_val)
                
                # Assign scores
                for score_col in F_SCORE_COLUMNS:
                    result_df.loc[mask, score_col] = scores[score_col]
                    
            except Exception as e:
                group_id = group if composite_key else "all"
                print(f"Error calculating F-Score for group {group_id} Year {current['Year']}: {str(e)}")
    
    # Clean up temporary column if it exists
    if '_temp_group_' in result_df.columns:
        result_df = result_df.drop(columns=['_temp_group_'])
    
    cols_to_drop = []

    # Remove index columns if they exist
    index_cols = [col for col in result_df.columns if col.startswith('Index')]
    cols_to_drop.extend(index_cols)

    # Drop all unwanted columns
    result_df = result_df.drop(columns=[col for col in cols_to_drop if col in result_df.columns])
    

    return result_df




def process_workbook(input_path: str, sheet_name: str = None, 
                    overwrite_policy: str = None, datetime_cols: Dict = None,
                    index_cols: List = None, file_type: str = 'excel',composite_key: List = None) -> bool:
    """
    Process Excel workbook to add F-Scores
    """
    try:
        book = load_workbook(input_path)
        sheets_to_process = [sheet_name] if sheet_name else book.sheetnames
        
        for sheet in sheets_to_process:
            print(f"\nProcessing sheet: {sheet}")
            
            # Read and standardize data
            df = read_data_with_custom_index(
                input_path, sheet_name=sheet,
                datetime_cols=datetime_cols,
                index_cols=index_cols,
                file_type=file_type
            )

            # Read and standardize data
            df = read_data_with_custom_index(
                input_path, sheet_name=sheet,
                datetime_cols=datetime_cols,
                index_cols=index_cols,
                file_type=file_type,
                composite_key=composite_key
            )

            # Handle existing F-Score columns
            existing_cols = [col for col in F_SCORE_COLUMNS if col in df.columns]
            if existing_cols:
                if overwrite_policy == 'ask':
                    response = input(f"Overwrite existing F-Score columns in {sheet}? (y/n): ").strip().lower()
                    overwrite = response == 'y'
                else:
                    overwrite = overwrite_policy == 'always'
                
                if not overwrite:
                    print(f"Skipping {sheet} - existing F-Score columns found")
                    continue
                
                df = df.drop(columns=existing_cols, errors='ignore')
            
            # Calculate F-Scores with composite key
            scored_df = calculate_f_scores(df, composite_key=composite_key)


            # Remove any Unnamed columns (like Unnamed: 0 or Unnamed: 0.1)
            for unwantedCol in ['Unnamed','Index']:
                unnamed_cols = [col for col in scored_df.columns if str(col).startswith(unwantedCol)]
                if unnamed_cols:
                    scored_df = scored_df.drop(columns=unnamed_cols, errors='ignore')
                    print(f"Removed unnamed columns: {unnamed_cols}")

            ws = book[sheet]
            
            # Clear existing data (keep headers)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            
            # Handle MultiIndex (Year and Quarter)
            if isinstance(scored_df.index, pd.MultiIndex):
                # Convert MultiIndex to columns
                scored_df.reset_index(inplace=True)
                
                # Write headers
                for c_idx, col in enumerate(scored_df.columns, 1):
                    ws.cell(row=1, column=c_idx, value=col)
                
                # Write data
                for r_idx, row in enumerate(scored_df.itertuples(index=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
            else:
                # For single index
                # Write headers
                for c_idx, col in enumerate(scored_df.columns, 2):
                    ws.cell(row=1, column=c_idx, value=col)
                
                # Write data with index in first column
                for r_idx, (index, row) in enumerate(scored_df.iterrows(), 2):
                    ws.cell(row=r_idx, column=1, value=index)
                    for c_idx, value in enumerate(row, 2):
                        ws.cell(row=r_idx, column=c_idx, value=value)
            
                # Add new headers if needed
                new_cols = [col for col in scored_df.columns if col not in df.columns]
                for c_idx, col in enumerate(new_cols, len(df.columns) + 2):
                    ws.cell(row=1, column=c_idx, value=col)
            
            print(f"Added F-Scores to {sheet}")
        
        book.save(input_path)
        return True
    
    except Exception as e:
        print(f"Error processing workbook: {str(e)}")
        return False

def main():
    """Interactive user interface for F-Score calculation"""
    print("=== Piotroski F-Score Calculator ===")
    
    # Get input file path
    
    while True:
        input_file = input("\nEnter path to Excel file: ").strip()
        if not input_file:
            print("Error: Please provide a file path")
            input_file = '/Users/ankit/Desktop/GitHub/AlgoTrading/QuantStrategies/TradingBot/data/output/tejori_financial_results.xlsx'
            break
            
        if not os.path.exists(input_file):
            print(f"Error: File not found - {input_file}")
            continue
            
        if not (input_file.endswith('.xlsx') or input_file.endswith('.xls')):
            print("Error: Only Excel files (.xlsx, .xls) are supported")
            continue
            
        break
    
    # Get sheet name (optional)
    sheet_name = input(
        "Enter specific sheet name to process (leave blank for all sheets): "
    ).strip() or None
    
    # Get composite key (new feature)
    composite_key = input(
        "Enter column names for composite key (comma-separated, leave blank if none): "
    ).strip()
    composite_key = [x.strip() for x in composite_key.split(',')] if composite_key else None
    if not composite_key:
        composite_key = ['tradingsymbol','Quarter']

    # Get processing options
    overwrite_policy = input(
        "Overwrite policy for existing F-Score columns?\n"
        "1. Ask for each sheet\n"
        "2. Always overwrite\n"
        "3. Never overwrite\n"
        "Choice (1-3): "
    ).strip()
    
    overwrite_policy = {
        '1': 'ask',
        '2': 'always',
        '3': 'never'
    }.get(overwrite_policy, 'always')
    
    # Process the workbook
    success = process_workbook(
        input_path=input_file,
        sheet_name=sheet_name,
        overwrite_policy=overwrite_policy,
        composite_key=composite_key
    )
    
    if success:
        print("\nProcessing completed successfully!")
        sys.exit(0)
    else:
        print("\nProcessing completed with errors")
        sys.exit(1)

if __name__ == "__main__":
    main()